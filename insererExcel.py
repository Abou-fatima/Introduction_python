import pandas as pd
from openpyxl import load_workbook

# Exemple de données à ajouter

a = input("Entrer votre nom : ")
b = input("Entrer votre âge : ")
c = input("Entrer votre ville : ")

data_to_add ={
    'Nom':[a],
    'Âge':[b],
    'Ville':[c] 
    }



"""
data_to_add = {
    'Nom': ['Diallo', 'Bah','Barry','Diaby'],
    'Âge': [28, 33,25,20],
    'Ville': ['Conakry', 'Mamou','Dalaba','Dabola']
}
"""


# Création d'un DataFrame à partir des données à ajouter
df_to_add = pd.DataFrame(data_to_add)

# Spécifiez le chemin du fichier Excel existant
file_path = 'C:/TPPYTHON/donnees_existantes.xlsx'

# Charger le fichier Excel existant
try:
    book = load_workbook(file_path)
    print("Fichier chargé avec succès.")
except FileNotFoundError:
    print(f"Erreur : le fichier {file_path} n'existe pas.")
    exit()

# Vérifier si la feuille existe, sinon créer une nouvelle feuille
sheet_name = 'Feuil1'  # Remplacez par le nom de votre feuille

# Si la feuille existe déjà, on la sélectionne
if sheet_name in book.sheetnames:
    sheet = book[sheet_name]
else:
    # Si la feuille n'existe pas, on la crée
    sheet = book.create_sheet(sheet_name)

# Ajouter les en-têtes si la feuille est vide
if sheet.max_row == 1:
    for col_num, column_title in enumerate(df_to_add.columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

# Ajouter les données à la feuille existante après la dernière ligne
for row in df_to_add.itertuples(index=False, name=None):
    sheet.append(row)

# Sauvegarder le fichier Excel avec les nouvelles données
book.save(file_path)
print(f"Les nouvelles données ont été ajoutées et le fichier a été sauvegardé dans {file_path}.")