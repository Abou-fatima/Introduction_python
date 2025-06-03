# -*- coding: utf-8 -*-
"""
Created on Wed May  7 14:35:07 2025

@author: abouf
"""

import tkinter as tk
import pandas as pd
from tkinter import messagebox
from openpyxl import load_workbook

def MettreSourrisSur1(event):
    BouttonAjouter['background']='green'
def EnleverSourrisSur1(event):
    BouttonAjouter['background']='white'
def MettreSourrisSur2(event):
    BouttonModifier['background']='green'
def EnleverSourrisSur2(event):
    BouttonModifier['background']='white'
def MettreSourrisSur3(event):
    BouttonSupprimer['background']='red'
def EnleverSourrisSur3(event):
    BouttonSupprimer['background']='white'
def MettreSourrisSur4(event):
    BouttonVider['background']='yellow'
def EnleverSourrisSur4(event):
    BouttonVider['background']='white'
def MettreSourrisSur5(event):
    BouttonRechercher['background']='green'
def EnleverSourrisSur5(event):
    BouttonRechercher['background']='white'

def ViderLesChamps():
    EntryMatricule.delete(0, tk.END)
    EntryNomPrenom.delete(0, tk.END)
    EntryAge.delete(0, tk.END)
    EntryVille.delete(0, tk.END)
    EntryRechercher.delete(0, tk.END)
    
def Ajouter():
    Matricule  =  EntryMatricule.get()    
    NomPrenom  =  EntryNomPrenom.get()
    Age        =  EntryAge.get()
    Ville      =  EntryVille.get
    
    if not Matricule or not NomPrenom or not Age or not Ville:
        messagebox.showerror("Erreur","Veuillez remplir tous les champs")
        return
    
    if not Age.isdigit():
        messagebox.showerror("Erreur","L'age doit être un entier")
        return
    
    Tableau_Excel = {
        "Matricule" : [Matricule],
        "Nom et Prenom" : [NomPrenom],
        "Age" : [Age],
        "Ville" : [Ville]
        }
    Ajouter_a_Excel = pd.DataFrame(Tableau_Excel)
    chemin_du_fichier = "C:/TPPYTHON/evaluation.xlsx"
    
    
    try:
        charger_fichier = load_workbook(chemin_du_fichier)
        print("Fichier chargé avec succès")
    except FileNotFoundError():
        print("Fichier non trouvé!!")
        
    nom_du_feuille = "Feuil1"
    if nom_du_feuille in charger_fichier.sheetnames:
        feuille = charger_fichier[nom_du_feuille]
    else:
        feuille = charger_fichier.create_sheet(nom_du_feuille)
        
    
    if feuille.max_row == 1:
        for numero_colonne, titre_colonne in enumerate(Ajouter_a_Excel.columns, 1):
            feuille.cell(row=1, column=numero_colonne, value=titre_colonne)
            
    for row in Ajouter_a_Excel.itertuples(index=False, name=None):
        feuille.append(row)
        
    charger_fichier.save(chemin_du_fichier)
    messagebox.showinfo("Info","Enregistrement effectuée avec succès")
    ViderLesChamps()
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
    
    
fenetre = tk.Tk()
fenetre.geometry("600x500+400+200")  # LongueurxLargeur + positionX + positionY
fenetre.title("Formulaire d'inscription dans excel")
fenetre['background'] = "blue"
fenetre.resizable(False, False)

labelMatricule = tk.Label(fenetre, text="Matricule", width=15)
labelMatricule.place(x=10, y=20)

labelNomPrenom = tk.Label(fenetre, text="Nom et Prenom", width=15)
labelNomPrenom.place(x=10, y=70)

labelAge = tk.Label(fenetre, text="Age", width=15)
labelAge.place(x=10, y=120)

labelVille = tk.Label(fenetre, text="Ville", width=15)
labelVille.place(x=10, y=170)

LabelRecherche = tk.Label(fenetre, text="Rechercher par Matricule", width=20)
LabelRecherche.place(x=290, y=120)

EntryMatricule = tk.Entry(fenetre)
EntryMatricule.place(x=140, y=20)


EntryNomPrenom = tk.Entry(fenetre)
EntryNomPrenom.place(x=140, y=70)


EntryAge = tk.Entry(fenetre)
EntryAge.place(x=140, y=120)


EntryVille = tk.Entry(fenetre)
EntryVille.place(x=140, y=170)

EntryRechercher = tk.Entry(fenetre)
EntryRechercher.place(x=440, y=120)



BouttonAjouter = tk.Button(fenetre, text="Ajouter", width=15, command=Ajouter)
BouttonAjouter.place(x=10, y=210)
BouttonAjouter.bind("<Enter>", MettreSourrisSur1)
BouttonAjouter.bind("<Leave>", EnleverSourrisSur1)

BouttonModifier = tk.Button(fenetre, text="Modifier", width=15)
BouttonModifier.place(x=150, y=210)
BouttonModifier.bind("<Enter>", MettreSourrisSur2)
BouttonModifier.bind("<Leave>", EnleverSourrisSur2)

BouttonSupprimer = tk.Button(fenetre, text="Supprimer", width=15)
BouttonSupprimer.place(x=290, y=210)
BouttonSupprimer.bind("<Enter>", MettreSourrisSur3)
BouttonSupprimer.bind("<Leave>", EnleverSourrisSur3)

BouttonVider = tk.Button(fenetre, text="Vider les champs", width=15, command=ViderLesChamps)
BouttonVider.place(x=440, y=210)
BouttonVider.bind("<Enter>", MettreSourrisSur4)
BouttonVider.bind("<Leave>", EnleverSourrisSur4)

BouttonRechercher = tk.Button(fenetre, text="Rechercher", width=16)
BouttonRechercher.place(x=440, y=150)
BouttonRechercher.bind("<Enter>", MettreSourrisSur5)
BouttonRechercher.bind("<Leave>", EnleverSourrisSur5)

fenetre.mainloop()
































