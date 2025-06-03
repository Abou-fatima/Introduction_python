import tkinter as tk
from tkinter import Menu
import os
from tkinter import ttk, messagebox
from openpyxl import load_workbook
import pandas as pd
from database import creer_tables
import mysql.connector
from tkcalendar import DateEntry
from database import connecter_bd 
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.utils import ImageReader
creer_tables()

# Fonctions pour ouvrir les fenêtres
def ouvrir_fenetre_enregistrer_eleve():
    def connecter():
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="scolarite"
        )
    
    def exporter_en_excel():
        chemin_fichier = "C:/Users/abouf/.spyder-py3/liste_eleves.xlsx"
        try:
            conn = connecter()
            cursor = conn.cursor()
            cursor.execute("SELECT matricule, nom, prenom, age, ville, date_naissance FROM eleves")
            donnees = cursor.fetchall()
    
            # Création du fichier Excel
            classeur = openpyxl.Workbook()
            feuille = classeur.active
            feuille.title = "Élèves"
    
            # En-têtes
            en_tetes = ["Matricule", "Nom", "Prénom", "Âge", "Ville", "Date de Naissance"]
            feuille.append(en_tetes)
            for cell in feuille[1]:
                cell.font = Font(bold=True)
    
            # Ajout des données
            for ligne in donnees:
                feuille.append(ligne)
    
            # Sauvegarde du fichier
            classeur.save(chemin_fichier)
            messagebox.showinfo("Succès", f"Fichier Excel enregistré sous : {chemin_fichier}")
    
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'exportation : {e}")
        finally:
            cursor.close()
            conn.close()
            
    def exporter_vers_pdf():
        data = []
        for row_id in tree.get_children():
            row = tree.item(row_id)['values']
            data.append(row)
        
        if not data:
            messagebox.showwarning("Export PDF", "Aucune donnée à exporter.")
            return
    
        try:
            # Ajouter les en-têtes comme première ligne du tableau
            data.insert(0, colonnes)
    
            chemin_fichier = 'C:/Users/abouf/.spyder-py3/liste_eleves.pdf'
            doc = SimpleDocTemplate(chemin_fichier, pagesize=letter)
    
            # Préparer les styles
            styles = getSampleStyleSheet()
            titre1 = Paragraph("MINISTERE DE L'ENSEIGNEMENT SUPERIEUR ET DE LA RECHERCHE SCIENTIFIQUE", styles['Title'])
            titre2 = Paragraph("Liste des Élèves", styles['Title'])

            # Création du tableau
            table = Table(data)
    
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ])
            table.setStyle(style)
    
            # Ajouter titre + espace + tableau
            elements = [titre1, Spacer(1, 12), titre2, Spacer(1, 20), table]
            doc.build(elements)
    
            messagebox.showinfo("Export PDF", f"Export réussi vers {chemin_fichier}")
        
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export PDF : {e}")
            
    def ajouter_filigrane(canvas, doc):
        # Charger le logo
        try:
            logo_path = 'C:/Users/abouf/.spyder-py3/logoIst.jpg'  # Mets ici ton vrai chemin
            logo = ImageReader(logo_path)
            largeur_page, hauteur_page = letter
    
            canvas.saveState()
            canvas.translate(largeur_page/2, hauteur_page/2)
            canvas.rotate(45)
            canvas.setFillAlpha(0.8)  # Opacité faible
            canvas.drawImage(logo, -150, -150, width=300, height=300, mask='auto')
            canvas.restoreState()
        except Exception as e:
            print(f"Erreur logo : {e}")
    

        
    def Enter1(event):
        BouttonAjouter['background']="green"
    def Enter2(event):
        BouttonAjouter['background']="white"
    
    def Enter3(event):
        BouttonModifier['background']="green"
    def Enter4(event):
        BouttonModifier['background']="white"
    
    def Enter5(event):
        BouttonSupprimer['background']="red"
    def Enter6(event):
        BouttonSupprimer['background']="white"
    
    def Enter7(event):
        BouttonRechercherEleve['background']="yellow"
    def Enter8(event):
        BouttonRechercherEleve['background']="white" 
        
        
    def charger_donnees_selectionnees(event):
        selected_item = tree.focus()
        if not selected_item:
            return
    
        valeurs = tree.item(selected_item, 'values')
        if valeurs:
            EntryMatricule.delete(0, tk.END)
            EntryMatricule.insert(0, valeurs[0])
    
            EntryNom.delete(0, tk.END)
            EntryNom.insert(0, valeurs[1])
    
            EntryPrenom.delete(0, tk.END)
            EntryPrenom.insert(0, valeurs[2])
    
            EntryAge.delete(0, tk.END)
            EntryAge.insert(0, valeurs[3])
    
            EntryVille.delete(0, tk.END)
            EntryVille.insert(0, valeurs[4])
    
            EntryDateNaiss.delete(0, tk.END)
            EntryDateNaiss.insert(0, valeurs[5])

    #Commencez ici les fonctions crud
        
    def afficher_eleves():
        
        for row in tree.get_children():
            tree.delete(row)
        conn = connecter_bd()
        cursor = conn.cursor()
        cursor.execute("SELECT matricule, nom, prenom, age, ville, date_naissance FROM eleves")
        rows = cursor.fetchall()
        for row in rows:
            tree.insert("", tk.END, values=row)
        cursor.close()
        conn.close()
    def charger_liste_eleves():
        
        try:
            EntryRechercherEleve.delete(0, tk.END)
            conn = connecter()
            cursor = conn.cursor()
            cursor.execute("SELECT matricule, nom, prenom, age, ville, date_naissance FROM eleves")
            lignes = cursor.fetchall()
    
            # Vider la Treeview
            for item in tree.get_children():
                tree.delete(item)
    
            # Recharger tous les élèves
            for ligne in lignes:
                tree.insert('', 'end', values=ligne)
    
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement : {e}")
        finally:
            cursor.close()
            conn.close()
    def vider_champs():
        EntryMatricule.delete(0, tk.END)
        EntryNom.delete(0, tk.END)
        EntryPrenom.delete(0, tk.END)
        EntryAge.delete(0, tk.END)
        EntryVille.delete(0, tk.END)
        EntryDateNaiss.delete(0, tk.END)
        EntryRechercherEleve.delete(0, tk.END)

    def ajouter_eleve():
        matricule = EntryMatricule.get().strip()
        nom = EntryNom.get().strip()
        prenom = EntryPrenom.get().strip()
        age = EntryAge.get().strip()
        ville = EntryVille.get().strip()
        date_naissance = EntryDateNaiss.get().strip()
    
        if not (matricule and nom and prenom and age and ville and date_naissance):
            messagebox.showwarning("Champs vides", "Tous les champs sont obligatoires.")
            return
    
        try:
            matricule = int(matricule)
            age = int(age)
        except ValueError:
            messagebox.showerror("Type invalide", "Le matricule et l'âge doivent être des nombres entiers.")
            return
    
        conn = connecter()
        cursor = conn.cursor()
        try:
            sql = "INSERT INTO eleves (matricule, nom, prenom, age, ville, date_naissance) VALUES (%s, %s, %s, %s, %s, %s)"
            valeurs = (matricule, nom, prenom, age, ville, date_naissance)
            cursor.execute(sql, valeurs)
            conn.commit()
            messagebox.showinfo("Succès", "Élève ajouté avec succès.")
            vider_champs()
            charger_liste_eleves()
            afficher_eleves()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout : {e}")
        finally:
            cursor.close()
            conn.close()

    def modifier_eleve():
           matricule = EntryMatricule.get().strip()
           nom = EntryNom.get().strip()
           prenom = EntryPrenom.get().strip()
           age = EntryAge.get().strip()
           ville = EntryVille.get().strip()
           date_naissance = EntryDateNaiss.get().strip()
       
           # Vérifications basiques
           if not (matricule and nom and prenom and age and ville and date_naissance):
               messagebox.showwarning("Champs vides", "Tous les champs sont obligatoires.")
               return
       
           try:
               matricule = int(matricule)
               age = int(age)
           except ValueError:
               messagebox.showerror("Type invalide", "Le matricule et l'âge doivent être des nombres entiers.")
               return
       
           conn = connecter()
           cursor = conn.cursor()
       
           try:
               cursor.execute("SELECT * FROM eleves WHERE matricule = %s", (matricule,))
               if cursor.fetchone() is None:
                   messagebox.showerror("Erreur", "Élève non trouvé avec ce matricule.")
                   return
               
               sql = """
               UPDATE eleves
               SET nom = %s, prenom = %s, age = %s, ville = %s, date_naissance = %s
               WHERE matricule = %s
               """
               valeurs = (nom, prenom, age, ville, date_naissance, matricule)
               cursor.execute(sql, valeurs)
               conn.commit()
               messagebox.showinfo("Succès", "Élève modifié avec succès.")
               vider_champs()
               afficher_eleves()
           except Exception as e:
               messagebox.showerror("Erreur", f"Erreur lors de la modification : {e}")
           finally:
               cursor.close()
               conn.close()
        
    def supprimer_eleve():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Attention", "Veuillez sélectionner un élève à supprimer.")
            return
    
        # Récupération du matricule (1ère colonne de la ligne sélectionnée)
        valeurs = tree.item(selected_item)['values']
        matricule = valeurs[0]
    
        confirmation = messagebox.askyesno("Confirmation", f"Voulez-vous vraiment supprimer l'élève avec le matricule {matricule} ?")
        if confirmation:
            try:
                conn = connecter()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM eleves WHERE matricule = %s", (matricule,))
                conn.commit()
                messagebox.showinfo("Succès", "Élève supprimé avec succès.")
                vider_champs()
                afficher_eleves()
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la suppression : {e}")
            finally:
                cursor.close()
                conn.close()
        
    def rechercher_eleves():
        terme = EntryRechercherEleve.get().strip()
        if not terme:
            messagebox.showwarning("Champ vide", "Veuillez entrer un terme de recherche.")
            return
    
        try:
            conn = connecter()
            cursor = conn.cursor()
            requete = """
                SELECT matricule, nom, prenom, age, ville, date_naissance
                FROM eleves
                WHERE matricule LIKE %s OR nom LIKE %s OR prenom LIKE %s
            """
            filtre = f"%{terme}%"
            cursor.execute(requete, (filtre, filtre, filtre))
            resultats = cursor.fetchall()
    
            for item in tree.get_children():
                tree.delete(item)
    
            for eleve in resultats:
                tree.insert('', 'end', values=eleve)
    
            if not resultats:
                messagebox.showinfo("Aucun résultat", "Aucun élève trouvé pour ce critère.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la recherche : {e}")
        finally:
            cursor.close()
            conn.close()

        
        
      
    fenetre = tk.Tk()
    fenetre.geometry("600x500+400+200")  # LongueurxLargeur + positionX + positionY
    fenetre.title("INSCRIPTION DES ELEVES")
    fenetre['background'] = "blue"
    fenetre.resizable(False, False)
    
    
    labelMatricule = tk.Label(fenetre, text="Matricule", width=15, fg="red", font=("Arial",10))
    labelMatricule.grid(row=0, column=0,padx =10, pady =10)
    EntryMatricule = tk.Entry(fenetre)
    EntryMatricule.grid(row=0, column=1, padx=10, pady=10 )
    
    labelNom= tk.Label(fenetre, text="Nom", width=15, fg="red", font=("Arial",10))
    labelNom.place(x=300, y=10)
    EntryNom = tk.Entry(fenetre)
    EntryNom.place(x=450, y=10)
    
    labelPrenom = tk.Label(fenetre, text="Prenom", width=15, fg="red", font=("Arial",10))
    labelPrenom.grid(row=1, column=0,padx =10, pady =10)
    EntryPrenom = tk.Entry(fenetre)
    EntryPrenom.place(x=155, y=53)
    
    labelAge = tk.Label(fenetre, text="Age", width=15, fg="red", font=("Arial",10))
    labelAge.place(x=300, y=50)
    EntryAge = tk.Entry(fenetre)
    EntryAge.place(x=450, y=53)
    
    labelVille = tk.Label(fenetre, text="Ville", width=15, fg="red", font=("Arial",10))
    labelVille.grid(row=2, column=0,padx =10, pady =10)
    EntryVille = tk.Entry(fenetre)
    EntryVille.grid(row=2, column=1, padx=10, pady=10 )
    
    labelDateNaiss = tk.Label(fenetre, text="Date naissance", width=15, fg="red", font=("Arial",10))
    labelDateNaiss.place(x=300, y=94)
    EntryDateNaiss = DateEntry(fenetre, width=17, date_pattern='yyyy-mm-dd')
    EntryDateNaiss.place(x=450, y=95)
    
    labelRecherche = tk.Label(fenetre, text="Rechercher", width=15)
    labelRecherche.place(x=330, y=150)
    EntryRechercherEleve = tk.Entry(fenetre)
    EntryRechercherEleve.place(x=450, y=150)   
    
    BouttonAjouter = tk.Button(fenetre, text="Ajouter", width=10, command=ajouter_eleve)
    BouttonAjouter.place(x=10, y=133)
    BouttonAjouter.bind("<Enter>",Enter1)
    BouttonAjouter.bind("<Leave>",Enter2)
    
    BouttonModifier = tk.Button(fenetre, text="Modifier",width=10, command=modifier_eleve)
    BouttonModifier.place(x=100, y=133)
    BouttonModifier.bind("<Enter>",Enter3)
    BouttonModifier.bind("<Leave>",Enter4)
    
    BouttonSupprimer = tk.Button(fenetre, text="Supprimer",width=10, command=supprimer_eleve)
    BouttonSupprimer.place(x=190, y=133)
    BouttonSupprimer.bind("<Enter>",Enter5)
    BouttonSupprimer.bind("<Leave>",Enter6)
    
    BouttonRechercherEleve = tk.Button(fenetre, text="Rechercher",width=10, command=rechercher_eleves)
    BouttonRechercherEleve.place(x=490, y=180)
    BouttonRechercherEleve.bind("<Enter>",Enter7)
    BouttonRechercherEleve.bind("<Leave>",Enter8)
    
    BouttonExporterPdf = tk.Button(fenetre, text="Exportez Vers PDF", width=15, command=exporter_vers_pdf)
    BouttonExporterPdf.place(x=10, y=220)
    
    BouttonExporterExcel = tk.Button(fenetre, text="Exportez Vers Excel", width=15, command=exporter_en_excel)
    BouttonExporterExcel.place(x=140, y=220)
    
    BouttonFiltrage = tk.Button(fenetre, text="Cliquez ici pour filtrer", width=18, command=ouvrir_fenetre_filtrage)
    BouttonFiltrage.place(x=270, y=220)
    
    
    btnActualiser = tk.Button(fenetre,width=10, text="Actualiser", command=charger_liste_eleves)
    btnActualiser.place(x=490, y=220)

    
    
    colonnes = ("Matricule", "Nom", "Prénom", "Age", "Ville", "Date Naissance")
    tree = ttk.Treeview(fenetre, columns=colonnes, show='headings')
    for col in colonnes:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    
    tree.place(x=0, y=300)
    tree.bind("<Double-1>", charger_donnees_selectionnees)

    afficher_eleves()
    
    fenetre.mainloop()
    
def ouvrir_fenetre_filtrage():
    import tkinter as tk
    from tkinter import messagebox
    from openpyxl import load_workbook
    import pandas as pd
    import re
    from datetime import datetime

    fenetre = tk.Tk()
    fenetre.geometry("620x400+400+200")  # LongueurxLargeur + positionX + positionY
    fenetre.title("FILTRAGE DES DONNEES")
    fenetre['background'] = "blue"
    fenetre.resizable(False, False)
    
    df = pd.read_excel("C:/Users/abouf/.spyder-py3/liste_eleves.xlsx")
    df_filtré = df.copy()

    ##################################################################
    def filtrer_avec_ia_avancee():
        nonlocal df_filtré
        requete = EntryFiltrage.get().lower()
        df_filtre = df.copy()
    
        # Filtres sur la ville
        villes_connues = ['mamou', 'kindia', 'conakry', 'labé']
        for ville in villes_connues:
            if ville in requete:
                df_filtre = df_filtre[df_filtre['Ville'].str.lower().str.contains(ville)]
    
        # Sexe
        if "fille" in requete or "féminin" in requete:
            df_filtre = df_filtre[df_filtre['Sexe'].str.lower().str.startswith('f')]
        elif "garçon" in requete or "masculin" in requete:
            df_filtre = df_filtre[df_filtre['Sexe'].str.lower().str.startswith('m')]
    
        # Nom commence par...
        match_nom = re.search(r"nom[s]? (qui )?(commence(nt)?|commençant) par (\w)", requete)
        if match_nom:
            lettre = match_nom.group(4).lower()
            df_filtre = df_filtre[df_filtre['Nom'].str.lower().str.startswith(lettre)]
    
        # Âge > ou < ou entre
        match_age = re.search(r"âge\s*(>|<|=|>=|<=)\s*(\d+)", requete)
        if match_age:
            operateur = match_age.group(1)
            age_limite = int(match_age.group(2))
            df_filtre['Age'] = df_filtre['DateNaissance'].apply(lambda d: datetime.now().year - int(d[:4]))
            df_filtre = df_filtre.query(f"Age {operateur} {age_limite}")
    
        df_filtré = df_filtre  # 🟢 Met à jour le DataFrame global filtré
    
        # Réinitialiser le Treeview
        for item in tree.get_children():
            tree.delete(item)
    
        # Afficher résultats
        for _, row in df_filtré.iterrows():
            tree.insert('', 'end', values=list(row.drop('Age', errors='ignore')))

 
    #################################################################################
    def actualiser():
        EntryFiltrage.delete(0, tk.END)
        afficher_donnees(df_filtré)

    EntryFiltrage = tk.Entry(fenetre)
    EntryFiltrage.place(x=15, y=45, width=480, height=30)
    
    def filtrer():
        critere = EntryFiltrage.get().lower()
        for item in tree.get_children():
            tree.delete(item)
        for _, row in df.iterrows():
            if critere in str(row.values).lower():
                tree.insert('', 'end', values=list(row))
                
    BouttonFiltrer = tk.Button(fenetre, text="Filtrer", width=10, height=1, command=filtrer_avec_ia_avancee)
    BouttonFiltrer.place(x=500, y=45)
    
    # Treeview
    colonnes = ["Matricule", "Nom", "Prénom", "Age","Ville"]
    tree = ttk.Treeview(fenetre, columns=colonnes, show='headings')
    for col in colonnes:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    tree.place(x=0, y=190, width=600, height=300)
    
    try:
        df = pd.read_excel('C:/Users/abouf/.spyder-py3/liste_eleves.xlsx')
        for _, row in df.iterrows():
            tree.insert('', 'end', values=list(row))
    except Exception as e:
        messagebox.showerror("Erreur", f"Impossible de charger les données : {e}")
        df = pd.DataFrame(columns=colonnes)
    
    
    BouttonExportPdf = tk.Button(fenetre, text="Exportez en pdf", width=15)
    BouttonExportPdf.place(x=5, y=120)

    
    
    def afficher_donnees(df):
            for item in tree.get_children():
                tree.delete(item)
            for _, row in df.iterrows():
                tree.insert('', 'end', values=list(row))
    afficher_donnees(df_filtré)
    
    def filtrer_donnees():
        nonlocal df_filtré
        requete = EntryFiltrage.get().lower()

        if "mamou" in requete:
            df_filtré = df[df['ville'].str.lower() == 'mamou']
        elif "nés en 2000" in requete or "né en 2000" in requete:
            df_filtré = df[df['date_naissance'].str.contains("2000")]
        elif "age > 23" in requete or "âge > 23" in requete:
            df_filtré = df[df['age'] > 23]
        elif "nom commence par a" in requete:
            df_filtré = df[df['nom'].str.upper().str.startswith('A')]
        else:
            messagebox.showinfo("Filtrage", "Requête non reconnue.")
            return
    afficher_donnees(df_filtré)
    
    def exporter_excel():
        if df_filtré.empty:
            messagebox.showwarning("Excel", "Aucune donnée filtrée à exporter.")
            return
        chemin = 'C:/Users/abouf/.spyder-py3/eleves_filtres.xlsx'
        df_filtré.to_excel(chemin, index=False)
        messagebox.showinfo("Excel", f"Données filtrées exportées vers : {chemin}")
        
    
    BouttonExportExcel = tk.Button(fenetre, text="Exportez en fichier Excel", width=20, command=exporter_excel)
    BouttonExportExcel.place(x=150, y=120)
                
    BouttonActualiser = tk.Button(fenetre, text="Actualiser", command=actualiser)
    BouttonActualiser.place(x=500, y=120)
    
    

    

    
    fenetre.mainloop()


def ouvrir_fenetre_classe():
    import tkinter as tk
    from tkinter import messagebox
    from openpyxl import load_workbook
    import pandas as pd
    
    def connecter():
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="scolarite"
        )
    
    def Enter1(event):
        BouttonAjouter['background']="green"
    def Enter2(event):
        BouttonAjouter['background']="white"
    
    def Enter3(event):
        BouttonModifier['background']="green"
    def Enter4(event):
        BouttonModifier['background']="white"
    
    def Enter5(event):
        BouttonSupprimer['background']="red"
    def Enter6(event):
        BouttonSupprimer['background']="white"
    
    def Enter7(event):
        BouttonRechercherClasse['background']="yellow"
    def Enter8(event):
        BouttonRechercherClasse['background']="white" 
        
        #Commencez ici les fonctions crude
        
    def charger_classe_selectionnee(event):
        global ancien_nom_classe
        selected_item = treeClasse.focus()
        if not selected_item:
            return
    
        valeurs = treeClasse.item(selected_item, 'values')
        if valeurs:
            EntryNomClasse.delete(0, tk.END)
            EntryNomClasse.insert(0, valeurs[0])
            ancien_nom_classe = valeurs[0]  # ← stocke le nom original
    
            EntryAnnee.delete(0, tk.END)
            EntryAnnee.insert(0, valeurs[1])

            
    def ajouter_classe():
        nom_classe = EntryNomClasse.get().strip()
        annee = EntryAnnee.get().strip()
    
        if not (nom_classe and annee):
            messagebox.showwarning("Champs vides", "Tous les champs sont obligatoires.")
            return
    
        try:
            conn = connecter()
            cursor = conn.cursor()
            sql = "INSERT INTO classes (nom_classe, annee_scolaire) VALUES (%s, %s)"
            cursor.execute(sql, (nom_classe, annee))
            conn.commit()
            messagebox.showinfo("Succès", "Classe ajoutée avec succès.")
            vider_champs_classe()
            afficher_classes()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout : {e}")
        finally:
            cursor.close()
            conn.close()
            
    def modifier_classe():
        global ancien_nom_classe
        nom_classe = EntryNomClasse.get().strip()
        annee = EntryAnnee.get().strip()
    
        if not (nom_classe and annee):
            messagebox.showwarning("Champs vides", "Tous les champs sont obligatoires.")
            return
    
        try:
            conn = connecter()
            cursor = conn.cursor()
    
            # Vérifie si la classe avec l'ancien nom existe
            cursor.execute("SELECT * FROM classes WHERE nom_classe = %s", (ancien_nom_classe,))
            if cursor.fetchone() is None:
                messagebox.showerror("Erreur", "Classe non trouvée.")
                return
    
            # Met à jour avec le nouveau nom et année
            sql = "UPDATE classes SET nom_classe = %s, annee_scolaire = %s WHERE nom_classe = %s"
            cursor.execute(sql, (nom_classe, annee, ancien_nom_classe))
            conn.commit()
    
            messagebox.showinfo("Succès", "Classe modifiée avec succès.")
            vider_champs_classe()
            afficher_classes()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la modification : {e}")
        finally:
            cursor.close()
            conn.close()

            
    def supprimer_classe():
        selected_item = treeClasse.selection()
        if not selected_item:
            messagebox.showwarning("Attention", "Veuillez sélectionner une classe à supprimer.")
            return
    
        nom_classe = treeClasse.item(selected_item)['values'][0]
    
        confirmation = messagebox.askyesno("Confirmation", f"Voulez-vous vraiment supprimer la classe {nom_classe} ?")
        if confirmation:
            try:
                conn = connecter()
                cursor = conn.cursor()
                cursor.execute("DELETE FROM classes WHERE nom_classe = %s", (nom_classe,))
                conn.commit()
                messagebox.showinfo("Succès", "Classe supprimée avec succès.")
                vider_champs_classe()
                afficher_classes()
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la suppression : {e}")
            finally:
                cursor.close()
                conn.close()
                
    def rechercher_classes():
        terme = EntryRechercherClasse.get().strip()
        if not terme:
            messagebox.showwarning("Champ vide", "Veuillez entrer un terme de recherche.")
            return
    
        try:
            conn = connecter()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT nom_classe, annee_scolaire 
                FROM classes 
                WHERE nom_classe LIKE %s OR annee_scolaire LIKE %s
            """, (f"%{terme}%", f"%{terme}%"))
            resultats = cursor.fetchall()
    
            treeClasse.delete(*treeClasse.get_children())
    
            for ligne in resultats:
                treeClasse.insert('', 'end', values=ligne)
    
            if not resultats:
                messagebox.showinfo("Aucun résultat", "Aucune classe trouvée.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la recherche : {e}")
        finally:
            cursor.close()
            conn.close()
            
    def afficher_classes():
        try:
            EntryRechercherClasse.delete(0, tk.END)
            conn = connecter()
            cursor = conn.cursor()
            cursor.execute("SELECT nom_classe, annee_scolaire FROM classes")
            lignes = cursor.fetchall()
    
            treeClasse.delete(*treeClasse.get_children())
    
            for ligne in lignes:
                treeClasse.insert('', 'end', values=ligne)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'affichage : {e}")
        finally:
            cursor.close()
            conn.close()
            
    def vider_champs_classe():
        EntryNomClasse.delete(0, tk.END)
        EntryAnnee.delete(0, tk.END)
        EntryRechercherClasse.delete(0, tk.END)


    fenetre = tk.Tk()
    fenetre.geometry("600x350+400+200")  # LongueurxLargeur + positionX + positionY
    fenetre.title("ENREGISTREMENT DES CLASSES")
    fenetre['background'] = "blue"
    fenetre.resizable(False, False)
    
    
    labelNomClasse = tk.Label(fenetre, text="Nom de Classe", width=15, fg="red", font=("Arial",10))
    labelNomClasse.grid(row=0, column=0,padx =10, pady =10)
    EntryNomClasse = tk.Entry(fenetre)
    EntryNomClasse.grid(row=0, column=1, padx=10, pady=10 )
    
    labelAnnee= tk.Label(fenetre, text="Annee Scolaire", width=15, fg="red", font=("Arial",10))
    labelAnnee.place(x=300, y=10)
    EntryAnnee = DateEntry(fenetre, width=17, date_pattern='yyyy-mm-dd')
    EntryAnnee.place(x=450, y=10)
    
    labelRechercheClasse = tk.Label(fenetre, text="Rechercher", width=15)
    labelRechercheClasse.place(x=330, y=80)
    EntryRechercherClasse = tk.Entry(fenetre)
    EntryRechercherClasse.place(x=450, y=80)
    
    
    # Les bouttons
    BouttonAjouter = tk.Button(fenetre, text="Ajouter",width=10, command=ajouter_classe)
    BouttonAjouter.place(x=10, y=70)
    BouttonAjouter.bind("<Enter>",Enter1)
    BouttonAjouter.bind("<Leave>",Enter2)
    
    BouttonModifier = tk.Button(fenetre, text="Modifier",width=10, command=modifier_classe)
    BouttonModifier.place(x=100, y=70)
    BouttonModifier.bind("<Enter>",Enter3)
    BouttonModifier.bind("<Leave>",Enter4)
    
    BouttonSupprimer = tk.Button(fenetre, text="Supprimer",width=10, command=supprimer_classe)
    BouttonSupprimer.place(x=190, y=70)
    BouttonSupprimer.bind("<Enter>",Enter5)
    BouttonSupprimer.bind("<Leave>",Enter6)
    
    BouttonRechercherClasse = tk.Button(fenetre, text="Rechercher",width=10, command=rechercher_classes)
    BouttonRechercherClasse.place(x=490, y=120)
    BouttonRechercherClasse.bind("<Enter>",Enter7)
    BouttonRechercherClasse.bind("<Leave>",Enter8)
    
    btnActualiser = tk.Button(fenetre,width=10, text="Actualiser", command=afficher_classes)
    btnActualiser.place(x=490, y=220)
    
    
    
    colonnes_classe = ("Nom de la Classe", "Année Scolaire")
    treeClasse = ttk.Treeview(fenetre, columns=colonnes_classe, show='headings')
    
    for col in colonnes_classe:
        treeClasse.heading(col, text=col)
        treeClasse.column(col, width=150)

    
    treeClasse.place(x=0, y=200)
    treeClasse.bind("<Double-1>", charger_classe_selectionnee)

    afficher_classes()
    
    fenetre.mainloop()
    
    
def ouvrire_fenetre_inscription():
    
    import tkinter as tk
    from tkinter import messagebox
    from openpyxl import load_workbook
    import pandas as pd
    
    def Enter1(event):
        BouttonAjouter['background']="green"
    def Enter2(event):
        BouttonAjouter['background']="white"
    
    def Enter3(event):
        BouttonModifier['background']="green"
    def Enter4(event):
        BouttonModifier['background']="white"
    
    def Enter5(event):
        BouttonSupprimer['background']="red"
    def Enter6(event):
        BouttonSupprimer['background']="white"
    
    def Enter7(event):
        BouttonRechercherInscription['background']="yellow"
    def Enter8(event):
        BouttonRechercherInscription['background']="white" 
    
    def connecter():
        return mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="scolarite"
        )
    def vider_champs():
        EntryInscription.delete(0, tk.END)
        EntryInscriptionClasse.delete(0, tk.END)
        EntryDateInscription.delete(0, tk.END)

    
    def charger_eleves():
        conn = connecter()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nom, prenom FROM eleves")
        eleves = cursor.fetchall()
        conn.close()
        return [f"{e[0]} - {e[1]} {e[2]}" for e in eleves]
    def charger_classes():
        conn = connecter()
        cursor = conn.cursor()
        cursor.execute("SELECT id, nom_classe FROM classes")
        classes = cursor.fetchall()
        conn.close()
        return [f"{c[0]} - {c[1]}" for c in classes]

    def ajouter_inscription():
        try:
            id_eleve = EntryInscription.get().split(" - ")[0]
            id_classe = EntryInscriptionClasse.get().split(" - ")[0]
            date_inscription = EntryDateInscription.get()
    
            if not (id_eleve and id_classe and date_inscription):
                messagebox.showwarning("Champs vides", "Tous les champs sont obligatoires.")
                return
    
            conn = connecter()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO inscription (id_eleve, id_classe, date_inscription) VALUES (%s, %s, %s)",
                           (id_eleve, id_classe, date_inscription))
            conn.commit()
            messagebox.showinfo("Succès", "Inscription ajoutée avec succès.")
            vider_champs()
            afficher_inscriptions()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l’ajout : {e}")
        finally:
            cursor.close()
            conn.close()
            
    def charger_inscription_selectionnee(event):
        selected = treeInscriptions.focus()
        if not selected:
            return
        valeurs = treeInscriptions.item(selected, 'values')
        print("Valeurs récupérées :", valeurs)
    
        if valeurs:
            EntryInscription.set(valeurs[1]) 
            EntryInscriptionClasse.set(valeurs[2])
            EntryDateInscription.set_date(valeurs[3])
            
    
    fenetre = tk.Tk()
    fenetre.geometry("600x450+400+200")  # LongueurxLargeur + positionX + positionY
    fenetre.title("INSCRIPTION DES ELEVES")
    fenetre['background'] = "blue"
    fenetre.resizable(False, False)
            
    treeInscriptions = ttk.Treeview(fenetre, columns=("ID", "Nom de L'élève", "Nom de la classe", "Date Inscription"), show='headings')
    treeInscriptions.heading("ID", text="ID")  # Facultatif à masquer après
    treeInscriptions.column("ID", width=0, stretch=False)
    treeInscriptions.heading("Nom de L'élève", text="Nom de L'élève")
    treeInscriptions.heading("Nom de la classe", text="Nom de la classe")
    treeInscriptions.heading("Date Inscription", text="Date Inscription")
    treeInscriptions.bind("<Double-1>", charger_inscription_selectionnee)
    treeInscriptions.place(x=0, y=250)

    def afficher_inscriptions():
        for item in treeInscriptions.get_children():
            treeInscriptions.delete(item)
        EntryRechercherInscription.delete(0, tk.END)
        conn = connecter()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.id, CONCAT(e.nom, ' ', e.prenom), c.nom_classe, i.date_inscription
            FROM inscription i
            JOIN eleves e ON i.id_eleve = e.id
            JOIN classes c ON i.id_classe = c.id
        """)
        for row in cursor.fetchall():
            treeInscriptions.insert("", tk.END, values=row)
        conn.close()
        
    def existe_eleve(id_eleve):
        conn = connecter_bd()
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM eleves WHERE id = %s", (id_eleve,))
        resultat = cursor.fetchone()
        conn.close()
        return resultat is not None

    def existe_classe(id_classe):
        conn = connecter_bd()
        cursor = conn.cursor()
        cursor.execute("SELECT 1 FROM classes WHERE id = %s", (id_classe,))
        resultat = cursor.fetchone()
        conn.close()
        return resultat is not None

        
    

            
    def modifier_inscription():
        selected = treeInscriptions.focus()
        if not selected:
            messagebox.showwarning("Sélection", "Aucune inscription sélectionnée.")
            return
    
        valeurs = treeInscriptions.item(selected, 'values')
        id_inscription = valeurs[0]
    
        # Supposons que EntryInscription.get() renvoie "1 - Bah Alpha Ousmane"
        id_eleve = EntryInscription.get().split(" - ")[0]
        id_classe = EntryInscriptionClasse.get().split(" - ")[0]
        date = EntryDateInscription.get()
    
        # Vérification avant mise à jour
        if not existe_eleve(id_eleve):
            messagebox.showerror("Erreur", f"L'élève avec id {id_eleve} n'existe pas.")
            return
        if not existe_classe(id_classe):
            messagebox.showerror("Erreur", f"La classe avec id {id_classe} n'existe pas.")
            return
    
        conn = connecter_bd()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE inscription SET id_eleve = %s, id_classe = %s, date_inscription = %s WHERE id = %s",
            (id_eleve, id_classe, date, id_inscription)
        )
        conn.commit()
        conn.close()
    
        messagebox.showinfo("Succès", "Inscription modifiée avec succès.")
        vider_champs()
        afficher_inscriptions()
    


        
    def rechercher_inscription():
        mot_cle = EntryRechercherInscription.get().strip()
        conn = connecter()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT i.id, CONCAT(e.nom, ' ', e.prenom), c.nom_classe, i.date_inscription
            FROM inscription i
            JOIN eleves e ON i.id_eleve = e.id
            JOIN classes c ON i.id_classe = c.id
            WHERE e.nom LIKE %s OR c.nom_classe LIKE %s
        """, (f"%{mot_cle}%", f"%{mot_cle}%"))
    
        treeInscriptions.delete(*treeInscriptions.get_children())
        for row in cursor.fetchall():
            treeInscriptions.insert("", tk.END, values=row)
        conn.close()
        
    def supprimer_inscription():
        selected_item = treeInscriptions.focus()
        if not selected_item:
            messagebox.showwarning("Sélection requise", "Veuillez sélectionner une inscription à supprimer.")
            return
    
        values = treeInscriptions.item(selected_item, 'values')
        id_inscription = values[0]
    
        confirmation = messagebox.askyesno("Confirmation", "Voulez-vous vraiment supprimer cette inscription ?")
        if not confirmation:
            return
    
        try:
            conn = connecter()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM inscription WHERE id = %s", (id_inscription,))
            conn.commit()
            messagebox.showinfo("Succès", "Inscription supprimée avec succès.")
            vider_champs()
            afficher_inscriptions()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la suppression : {e}")
        finally:
            cursor.close()
            conn.close()




        
      #Commencez ici les fonctions crude
      
    
    
    labelInscription = tk.Label(fenetre, text="Eleve", width=15, fg="red", font=("Arial",10))
    labelInscription.grid(row=0, column=0,padx =10, pady =10)
    EntryInscription = ttk.Combobox(fenetre,values=charger_eleves(), width=25, state="readonly")
    EntryInscription.grid(row=0, column=1, padx=10, pady=10 )
    
    labelClasseInscription= tk.Label(fenetre, text="Classe", width=15, fg="red", font=("Arial",10))
    labelClasseInscription.place(x=340, y=10)
    EntryInscriptionClasse = ttk.Combobox(fenetre,values=charger_classes(), width=15, state="readonly")
    EntryInscriptionClasse.place(x=480, y=10)
    
    labelDateInscription = tk.Label(fenetre, text="Date inscription", width=15, fg="red", font=("Arial",10))
    labelDateInscription.grid(row=1, column=0,padx =10, pady =10)
    EntryDateInscription= DateEntry(fenetre, width=17, date_pattern='yyyy-mm-dd')
    EntryDateInscription.place(x=155, y=53)
    
    labelRecherche = tk.Label(fenetre, text="Rechercher", width=15)
    labelRecherche.place(x=330, y=150)
    EntryRechercherInscription = tk.Entry(fenetre)
    EntryRechercherInscription.place(x=450, y=150)
    
    
    # Les bouttons
    BouttonAjouter = tk.Button(fenetre, text="Ajouter",width=10, command=ajouter_inscription)
    BouttonAjouter.place(x=10, y=133)
    BouttonAjouter.bind("<Enter>",Enter1)
    BouttonAjouter.bind("<Leave>",Enter2)
    
    BouttonModifier = tk.Button(fenetre, text="Modifier",width=10, command=modifier_inscription)
    BouttonModifier.place(x=100, y=133)
    BouttonModifier.bind("<Enter>",Enter3)
    BouttonModifier.bind("<Leave>",Enter4)
    
    BouttonSupprimer = tk.Button(fenetre, text="Supprimer",width=10, command=supprimer_inscription)
    BouttonSupprimer.place(x=190, y=133)
    BouttonSupprimer.bind("<Enter>",Enter5)
    BouttonSupprimer.bind("<Leave>",Enter6)
    
    BouttonRechercherInscription = tk.Button(fenetre, text="Rechercher",width=10, command=rechercher_inscription)
    BouttonRechercherInscription.place(x=490, y=180)
    BouttonRechercherInscription.bind("<Enter>",Enter7)
    BouttonRechercherInscription.bind("<Leave>",Enter8)
    
    btnActualiser = tk.Button(fenetre,width=10, text="Actualiser", command=afficher_inscriptions)
    btnActualiser.place(x=490, y=220)
    

    
    afficher_inscriptions()
    
    fenetre.mainloop()


def ouvrire_fenetre_paiement():
    import tkinter as tk
    from tkinter import messagebox
    from openpyxl import load_workbook
    import pandas as pd
    
    def Enter1(event):
        BouttonAjouter['background']="green"
    def Enter2(event):
        BouttonAjouter['background']="white"
    
    def Enter3(event):
        BouttonModifier['background']="green"
    def Enter4(event):
        BouttonModifier['background']="white"
    
    def Enter5(event):
        BouttonSupprimer['background']="red"
    def Enter6(event):
        BouttonSupprimer['background']="white"
    
    def Enter7(event):
        BouttonRechercherPaiement['background']="yellow"
    def Enter8(event):
        BouttonRechercherPaiement['background']="white"  
        
        #Commencez ici les fonctions crude
        
            
    
    fenetre = tk.Tk()
    fenetre.geometry("600x450+400+200")  # LongueurxLargeur + positionX + positionY
    fenetre.title("PAIEMENT DES ELEVES")
    fenetre['background'] = "blue"
    fenetre.resizable(False, False)
    
    
    labelPaiementEleve = tk.Label(fenetre, text="Eleve", width=15, fg="red", font=("Arial",10))
    labelPaiementEleve.grid(row=0, column=0,padx =10, pady =10)
    EntryPaiementEleve = tk.Entry(fenetre)
    EntryPaiementEleve.grid(row=0, column=1, padx=10, pady=10 )
    
    labelPaiementClasse= tk.Label(fenetre, text="Classe", width=15, fg="red", font=("Arial",10))
    labelPaiementClasse.place(x=300, y=10)
    EntryPaiementClasse= tk.Entry(fenetre)
    EntryPaiementClasse.place(x=450, y=10)
    
    labelPaiementMontant = tk.Label(fenetre, text="Montant", width=15, fg="red", font=("Arial",10))
    labelPaiementMontant.grid(row=1, column=0,padx =10, pady =10)
    EntryPaiementMontant = tk.Entry(fenetre)
    EntryPaiementMontant.place(x=155, y=53)
    
    labelDatePaiement = tk.Label(fenetre, text="Date Paiement", width=15, fg="red", font=("Arial",10))
    labelDatePaiement.place(x=300, y=50)
    EntryDatePaiemen = tk.Entry(fenetre)
    EntryDatePaiemen.place(x=450, y=53)
    
    labelRecherche = tk.Label(fenetre, text="Rechercher", width=15)
    labelRecherche.place(x=330, y=150)
    EntryRechercher = tk.Entry(fenetre)
    EntryRechercher.place(x=450, y=150)
    
    
    # Les bouttons
    BouttonAjouter = tk.Button(fenetre, text="Ajouter",width=10)
    BouttonAjouter.place(x=10, y=110)
    BouttonAjouter.bind("<Enter>",Enter1)
    BouttonAjouter.bind("<Leave>",Enter2)
    
    BouttonModifier = tk.Button(fenetre, text="Modifier",width=10)
    BouttonModifier.place(x=100, y=110)
    BouttonModifier.bind("<Enter>",Enter3)
    BouttonModifier.bind("<Leave>",Enter4)
    
    BouttonSupprimer = tk.Button(fenetre, text="Supprimer",width=10)
    BouttonSupprimer.place(x=190, y=110)
    BouttonSupprimer.bind("<Enter>",Enter5)
    BouttonSupprimer.bind("<Leave>",Enter6)
    
    BouttonRechercherPaiement = tk.Button(fenetre, text="Rechercher",width=10)
    BouttonRechercherPaiement.place(x=490, y=110)
    BouttonRechercherPaiement.bind("<Enter>",Enter7)
    BouttonRechercherPaiement.bind("<Leave>",Enter8)
    
    fenetre.mainloop()



# Création de la fenêtre principale
root = tk.Tk()
root.title("GESTION DE SCOLARITE DES ELEVES")
root.geometry("650x500+400+200")
root['background']='white'
root.resizable(False, False)
 
menubar = Menu(root)

menu_gestion_etudiant = Menu(menubar, tearoff=0)
menu_gestion_etudiant.add_command(label="Inscrire un eleve", command=ouvrir_fenetre_enregistrer_eleve)
#menu_gestion_etudiant.add_command(label="Liste des eleves")
menubar.add_cascade(label="Gestion des Eleves", menu=menu_gestion_etudiant)

menu_gestion_listes = Menu(menubar, tearoff=0)
menu_gestion_listes.add_command(label="Enregistrer une classe", command=ouvrir_fenetre_classe)
menubar.add_cascade(label="Gestion des Classes", menu=menu_gestion_listes)

menu_gestion_inscription = Menu(menubar, tearoff=0)
menu_gestion_inscription.add_command(label="Faire une inscription", command=ouvrire_fenetre_inscription)
menubar.add_cascade(label="Gestion des inscriptions", menu=menu_gestion_inscription)

menu_gestion_paiement = Menu(menubar, tearoff=0)
menu_gestion_paiement.add_command(label="Efectuer un paiement",command=ouvrire_fenetre_paiement)
menubar.add_cascade(label="Gestion des Paiements", menu=menu_gestion_paiement)

menu_gestion_parametre = Menu(menubar, tearoff=0)
menu_gestion_parametre.add_command(label="Deconnexion")
menubar.add_cascade(label="Parametre", menu=menu_gestion_parametre)

# Affectation de la barre de menu à la fenêtre
root.config(menu=menubar)

# Boucle principale
root.mainloop()
