 # -*- coding: utf-8 -*-
"""
Created on Wed Apr  9 09:46:44 2025

@author: abouf
"""

import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import pandas as pd

def Enter1(event):
    BouttonAjouter['background']="green"
def Enter2(event):
    BouttonAjouter['background']="white"

def Enter3(event):
    BouttonModifier['background']="green"
def Enter4(event):
    BouttonModifier['background']="white"

def Enter5(event):
    BouttonSupprimer['background']="red"
def Enter6(event):
    BouttonSupprimer['background']="white"

def Enter7(event):
    BouttonRechercher['background']="yellow"
def Enter8(event):
    BouttonRechercher['background']="white"    
    

# la fonction qui recupéres et insère les données dans la base de données
def AjouterDonnees():
        Nom = EntryNom.get()
        Age = EntryAge.get()
        Ville = EntryVille.get()
        
        # Verifier que tous les champs soient remoplis*
        if not Nom or not Age or not Ville:
            messagebox.showerror('erreur','Tous les champs sont obligatoirs')
            return
        
        # verifier si l'age est un entier
        if not Age.isdigit():
               messagebox.showerror('erreur','l\'age doit être un nombre')
               return
  
                          
        data_to_add ={
          'Nom':[Nom],
          'Matricule':[Age],
          'Ville':[Ville] 
          }
        # Création d'un DataFrame à partir des données à ajouter
        df_to_add = pd.DataFrame(data_to_add)
        
        # Spécifiez le chemin du fichier Excel existant
        file_path = 'C:/TPPYTHON/donnees_existantes.xlsx'
        
        # Charger le fichier Excel existant
        try:
            book = load_workbook(file_path)
            print("Fichier chargé avec succès.")
        except FileNotFoundError:
            print(f"Erreur : le fichier {file_path} n'existe pas.")
        
        # Vérifier si la feuille existe, sinon créer une nouvelle feuille
        sheet_name = 'Feuil1' # Remplacez par le nom de votre feuille
        
        # Si la feuille existe déjà, on la sélectionne
        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
        else:
            # Si la feuille n'existe pas, on la crée
            sheet = book.create_sheet(sheet_name)
        
        # Ajouter les en-têtes si la feuille est vide
        if sheet.max_row == 1:
            for col_num, column_title in enumerate(df_to_add.columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
        
        # Ajouter les données à la feuille existante après la dernière ligne
        for row in df_to_add.itertuples(index=False, name=None):
            sheet.append(row)
        
        # Sauvegarder le fichier Excel avec les nouvelles données
        book.save(file_path)
        messagebox.showinfo("Succès","Insertion effectuée avec succès")
        EntryNom.delete(0, tk.END)
        EntryAge.delete(0, tk.END)
        EntryVille.delete(0, tk.END)
        print(f"Les nouvelles données ont été ajoutées et le fichier a été sauvegardé dans {file_path}.")
                
def rechercher():
    rechercherVal = EntryRechercher.get().strip()
    if not rechercherVal:
        messagebox.showerror('erreur','Aucune données ne coresspond')
        return
    data = pd.read_excel('C:/TPPYTHON/donnees_existantes.xlsx')
    dataRecherche = data[data['Matricule'].astype(str)==rechercherVal]
    # Recuperation des champs de filtrage
    if not dataRecherche.empty:
        nom1 = dataRecherche.iloc[0]['Nom']
        age1 =dataRecherche.iloc[0]['Matricule']
        ville1 = dataRecherche.iloc[0]['Ville']
        EntryNom.delete(0, tk.END)
        EntryNom.insert(0,nom1)
        EntryAge.delete(0, tk.END) 
        EntryAge.insert(0, age1)
        EntryVille.delete(0, tk.END)
        EntryVille.insert(0, ville1)
        messagebox.showinfo("Vous avez trouvé","Recherche effectué avec succès")
    else:
        messagebox.showerror("Pas trouvé","Echec de Recherche")
    global ligne_trouvee
    ligne_trouvee=dataRecherche.iloc[0]
def supprimer():
    if not ligne_trouvee.empty:
        reponse=messagebox.askyesno("Confirmation","Etes-vous sûr de vouloir supprimer")
        if reponse: 
            try:
                df=pd.read_excel("C:/TPPYTHON/donnees_existantes.xlsx")
                df=df[df['Matricule']!=ligne_trouvee['Matricule']]
                df.to_excel("C:/TPPYTHON/donnees_existantes.xlsx",index=False)
                messagebox.showinfo("Succès","Donnée supprimée avec succès.")
                EntryRechercher.delete(0,tk.END )  
                EntryNom.delete(0, tk.END)
                EntryAge.delete(0, tk.END)
                EntryVille.delete(0, tk.END)
                #Vider le champs rechercher après suppression
            except Exception:
                messagebox.showerror("Erreur",f"Erreur lors de la suppression des données")
        else:
            messagebox.showinfo("Annulation","Aucun résultat trouvé à supprimer")
    else:
        messagebox.showwarning("Attention","Aucun résultat trouvé à supprimer")
        
def modifier_donnees():
    if not ligne_trouvee.empty:
      
        nouveau_nom=EntryNom.get()
        nouveau_age=EntryAge.get()
        nouvelle_ville=EntryVille.get()
        if not nouveau_nom or not nouveau_age or not nouvelle_ville:
            messagebox.showinfo("Attention","Tous les champs sont obligatoire")
            return 
        try:
            df=pd.read_excel("C:/TPPYTHON/donnees_existantes.xlsx",sheet_name="Feuil1")
       
        
            df.loc[df["Nom"]==ligne_trouvee["Nom"],"Nom"]=nouveau_nom
            df.loc[df["Matricule"]==ligne_trouvee["Matricule"],"Matricule"]=nouveau_age
            df.loc[df["Ville"]==ligne_trouvee["Ville"],"Ville"]=nouvelle_ville
             
            df.to_excel("C:/TPPYTHON/donnees_existantes.xlsx",sheet_name="Feuil1",index=False)
            messagebox.showinfo("Success","Données modifiée avec succès")
            EntryAge.delete(0,tk.END)
            EntryNom.delete(0,tk.END)
            EntryRechercher.delete(0,tk.END)
            EntryVille.delete(0,tk.END)
        except Exception as e:
            messagebox.showerror("Erreur",f"Erreur lors de la modification: {str(e)}")
        else:
            messagebox.showarring("Attention","Aucun resultat trouvé à modifier")
  
        
        
        
        

fenetre = tk.Tk()
fenetre.geometry("600x500+400+200")  # LongueurxLargeur + positionX + positionY
fenetre.title("INSCRIPTION DES ELEVES")
fenetre['background'] = "blue"
fenetre.resizable(False, False)


labelMatricule = tk.Label(fenetre, text="Matricule", width=15, fg="red", font=("Arial",10))
labelMatricule.grid(row=0, column=0,padx =10, pady =10)
labelMatricule = tk.Entry(fenetre)
labelMatricule.grid(row=0, column=1, padx=10, pady=10 )

labelNom= tk.Label(fenetre, text="Nom", width=15, fg="red", font=("Arial",10))
labelNom.place(x=300, y=10)
EntryNom = tk.Entry(fenetre)
EntryNom.place(x=450, y=10)

labelPrenom = tk.Label(fenetre, text="Prenom", width=15, fg="red", font=("Arial",10))
labelPrenom.grid(row=1, column=0,padx =10, pady =10)
EntryPrenom = tk.Entry(fenetre)
EntryPrenom.place(x=155, y=53)

labelAge = tk.Label(text="Age", width=15, fg="red", font=("Arial",10))
labelAge.place(x=300, y=50)
EntryAge = tk.Entry(fenetre)
EntryAge.place(x=450, y=53)

labelVille = tk.Label(fenetre, text="Ville", width=15, fg="red", font=("Arial",10))
labelVille.grid(row=2, column=0,padx =10, pady =10)
EntryVille = tk.Entry(fenetre)
EntryVille.grid(row=2, column=1, padx=10, pady=10 )

labelDateNaiss = tk.Label(text="Date naissance", width=15, fg="red", font=("Arial",10))
labelDateNaiss.place(x=300, y=94)
EntryDateNaiss = tk.Entry(fenetre)
EntryDateNaiss.place(x=450, y=95)

labelRecherche = tk.Label(fenetre, text="Rechercher", width=15)
labelRecherche.place(x=330, y=150)
EntryRechercher = tk.Entry(fenetre)
EntryRechercher.place(x=450, y=150)


# Les bouttons
BouttonAjouter = tk.Button(fenetre, text="Ajouter",width=10, command=AjouterDonnees)
BouttonAjouter.place(x=10, y=133)
BouttonAjouter.bind("<Enter>",Enter1)
BouttonAjouter.bind("<Leave>",Enter2)

BouttonModifier = tk.Button(fenetre, text="Modifier",width=10, command=modifier_donnees)
BouttonModifier.place(x=100, y=133)
BouttonModifier.bind("<Enter>",Enter3)
BouttonModifier.bind("<Leave>",Enter4)

BouttonSupprimer = tk.Button(fenetre, text="Supprimer",width=10, command=supprimer)
BouttonSupprimer.place(x=190, y=133)
BouttonSupprimer.bind("<Enter>",Enter5)
BouttonSupprimer.bind("<Leave>",Enter6)

BouttonRechercher = tk.Button(fenetre, text="Rechercher",width=10, command=rechercher)
BouttonRechercher.place(x=490, y=180)
BouttonRechercher.bind("<Enter>",Enter7)
BouttonRechercher.bind("<Leave>",Enter8)

fenetre.mainloop()





































