# -*- coding: utf-8 -*-
"""
Created on Tue Apr  8 13:20:17 2025

@author: HP
"""
import pandas as pd
from openpyxl import load_workbook
def insertiondonnee():
    nom=saisi_nom.get()
    age=saisi_age.get()
    ville=saisi_ville.get()
    matricule_base=saisi_matricule.get()
    #Vérifier si tous les champs ne sont pas vides
    if not nom or not age or not ville or not matricule_base:
        messagebox.showerror("Erreurs les champs ne doivent pas être vide","champs obligatoire")
        return
    #Vérifier si l'age est un nombre
    if not age.isdigit():
        messagebox.showerror("Erreurs ,l'age doit être un nombre","champs obligatoire")
        return
    data_to_add={
        'Nom' :[nom],
        'Age':[age],
        'Ville':[ville],
        'Matricule':[matricule_base]
        }
    df_to_add=pd.DataFrame(data_to_add)
    print(df_to_add)
    file_path='C:/TPPYTHON/fenetre1.xlsx'
    try:
        book = load_workbook(file_path)
        print("Fichier chargé avec succès.")
    except FileNotFoundError:
        print(f"Erreur : le fichier {file_path} n'existe pas.")
        exit()
    #Vérifier si la feuille existe sinon créer
    sheet_name='Feuil1' #Remplacer par le nom de votre Feuille
    #Si elle existe déjà on selectionne
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        sheet = book.create_sheet(sheet_name)
    #Ajouter les entêtes si la feuille est vide
    if sheet.max_row:
        for col_num, column_title in enumerate(df_to_add.columns, 1):
            sheet.cell(row=1, column=col_num, value=column_title)
    for row in df_to_add.itertuples(index=False, name=None):
        sheet.append(row)
        book.save(file_path)
        messagebox.showinfo("Enregistrement","Enregistrement effectué avec succès") 
    # Sauvegarder le fichier Excel avec les nouvelles données
    book.save(file_path)
    print(f"Les nouvelles données ont été ajoutées et le fichier a été sauvegardé dans {file_path}.")
    saisi_nom.delete(0,tk.END)
    saisi_age.delete(0,tk.END)
    saisi_ville.delete(0,tk.END)
    saisi_matricule.delete(0,tk.END)
#la bibliothèque qui créer une fenêtre
import tkinter as tk 
#importer le message dans tkinter
from tkinter import messagebox
#la bibliothèque qui permet de charger les données dans excel
from openpyxl import load_workbook
#importer la widget dans laquelle se trouve la classe combobox
from tkinter import ttk
# Création de la fenêtre
fenetre=tk.Tk()
fenetre.title("Ajout Personnel")

#Lancement de la fênêtre
#fenetre.mainloop()
#Définition de la taille de la fenêtre et de sa position
#fenetre.geometry(300700)
#Création des champs de Labels
#Fonction pour récupérer la valeur selectionnée
def rechercher():
    recherchevaleur=saisi_rechercher.get()
    #Supprimer les espaces inutiles
    recherchevaleur=saisi_rechercher.get().strip()
    if not recherchevaleur:
        messagebox.showerror("Désolé aucun élément de correspond à votre recherche","Element non trouvé")
        return
    data=pd.read_excel("C:/TP_Python_Analyse_donnee/fichier_insertion_donnee.xlsx")
    #Récupérer la colonne Matricule et la mettre dans data_rechercher
    data_rechercher=data[data['Matricule'].astype(str)==recherchevaleur]
    #Récupérer la colonne Nom et la mettre dans data_rechercherNom
    data_rechercherNom=data[data['Nom'].astype(str)==recherchevaleur]
    #Récupérer la colonne Ville et la mettre dans data_rechercherVille
    data_rechercherVille=data[data['Ville'].astype(str)==recherchevaleur]
    #Récupérer la colonne Age et la mettre dans data_rechercherVille
    data_rechercherAge=data[data['Age'].astype(str)==recherchevaleur]
    #Contrôler les valeurs récupérées dans data_Rechercher
    if not data_rechercher.empty:
        matricule1=data_rechercher.iloc[0]['Matricule']
        saisi_matricule.delete(0,tk.END)
        saisi_matricule.insert(0,matricule1)
        nom1=data_rechercher.iloc[0]['Nom']
        saisi_nom.delete(0,tk.END)
        saisi_nom.insert(0,nom1)
        age1=data_rechercher.iloc[0]['Age']
        saisi_age.delete(0,tk.END)
        saisi_age.insert(0,age1)
        ville1=data_rechercher.iloc[0]['Ville']
        saisi_ville.delete(0,tk.END)
        saisi_ville.insert(0,ville1)
        messagebox.showinfo("Vous avez trouvé","Recherche effectué avec succès")
    else:
        messagebox.showerror("Pas trouvé","Echec de Recherche")
    global ligne_trouvee
    ligne_trouvee=data_rechercher.iloc[0]
def modifier_donnee():
    if not ligne_trouvee.empty:
        #Récupérer la nouvelle valeur à modifier
        nouvelle_valeur=saisi_rechercher.get().strip()
        if nouvelle_valeur=="":
            messagebox.showwarning("Attention","Veuillez entrer la valeur")
            return
        #Modifier la donnée dans le DataFrame
        try:
            df=pd.read_excel("C:/TP_Python_Analyse_donnee/fichier_insertion_donnee.xlsx")
            df.loc[df['Matricule']==ligne_trouvee['Matricule'],'Matricule']=nouvelle_valeur
            df.to_excel("C:/TP_Python_Analyse_donnee/fichier_insertion_donnee1.xlsx",index=False)
            messagebox.showinfo("Success","Donnée modifiée avec succès")
        except Exception as e:
            messagebox.showerror("Erreur",f"Erreur lors de la modification des données")
    else:
        messagebox.showwarning("Attention","Aucun resultat trouvé à modifer.")
def supprimer():
    if not ligne_trouvee.empty:
        #Demander une confirmation avant de supprimer
        reponse=messagebox.askyesno("Confirmation","Etes-vous sûr de vouloir supprimer")
        if reponse: 
            try:
                df=pd.read_excel("C:/TP_Python_Analyse_donnee/fichier_insertion_donnee.xlsx")
                df=df[df['Matricule']!=ligne_trouvee['Matricule']]
                df.to_excel("C:/TP_Python_Analyse_donnee/fichier_insertion_donnee.xlsx",index=False)
                messagebox.showinfo("Succès","Donnée supprimée avec succès.")
                saisi_rechercher.delete(0,tk.END )#Vider le champs rechercher après suppression
            except Exception:
                messagebox.showerror("Erreur",f"Erreur lors de la suppression des données")
        else:
            messagebox.showinfo("Annulation","Aucun résultat trouvé à supprimer")
    else:
        messagebox.showwarning("Attention","Aucun résultat trouvé à supprimer")

    #Fonction pourmodifier la donnée dans le fichier excel
#Création des labels
label_nom=tk.Label(fenetre,text="Nom")
label_age=tk.Label(fenetre,text="Age")
label_ville=tk.Label(fenetre,text="Ville")
label_nom.grid(row=0,column=0,padx=10,pady=5)
label_age.grid(row=1,column=0,padx=10,pady=5)
label_ville.grid(row=2,column=0,padx=10,pady=5)
label_matricule=tk.Label(fenetre,text="Matricule")
label_matricule.grid(row=3,column=0,padx=10,pady=5)
#Création des champs de saisies
saisi_nom=tk.Entry(fenetre)
saisi_nom.grid(row=0,column=1,padx=10,pady=5)
saisi_age=tk.Entry(fenetre)
saisi_age.grid(row=1,column=1,padx=10,pady=5)
saisi_ville=tk.Entry(fenetre)
saisi_ville.grid(row=2,column=1,padx=10,pady=5)
saisi_matricule=tk.Entry(fenetre)
saisi_matricule.grid(row=3,column=1,padx=10,pady=5)
saisi_rechercher=tk.Entry(fenetre)
saisi_rechercher.grid(row=4,column=3,padx=10,pady=10)
#Créer un bouton pour ajouter les données
bouton_Ajouter=tk.Button(fenetre,text="AJOUTER",command=insertiondonnee)
bouton_Ajouter.grid(row=4,column=0,padx=10,pady=5)
bouton_rechercher=tk.Button(fenetre,text="Rechercher",command=rechercher)
bouton_rechercher.grid(row=4,column=1,padx=10,pady=5)
bouton_modifier=tk.Button(fenetre,text="Modifier",command=modifier_donnee)
bouton_modifier.grid(row=4,column=4,padx=10,pady=10)
bouton_supprimer=tk.Button(fenetre,text="Supprimer",command=supprimer)
bouton_supprimer.grid(row=4,column=5,padx=10,pady=10)
#Création d'une combobox
options=["Nom","Age","ville","Matricule"]
combo=tk.ttk.Combobox(fenetre,values=options)
combo.grid(row=4,column=2,padx=10,pady=10)
#définir une valeur par défaut
combo.set("Nom")
fenetre.mainloop()



