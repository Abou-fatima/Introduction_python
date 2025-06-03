# -*- coding: utf-8 -*-
"""
Ajout manuel de données dans un fichier Excel avec vérification de doublons.
"""
import pandas as pd
from openpyxl import load_workbook

# Spécifiez le chemin du fichier Excel existant
file_path = 'C:/TPPYTHON/donnees_existantes.xlsx'

# Charger le fichier Excel existant
try:
    book = load_workbook(file_path)
    print("Fichier chargé avec succès.")
except FileNotFoundError:
    print(f"Erreur : le fichier {file_path} n'existe pas.")
    exit()

# Vérifier si la feuille existe, sinon créer une nouvelle feuille
sheet_name = 'Feuil1'
if sheet_name in book.sheetnames:
    sheet = book[sheet_name]
else:
    sheet = book.create_sheet(sheet_name)

# Lire les combinaisons existantes (Nom, Âge, Ville)
existing_combinations = set()
for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
    if all(row):  # S'assurer que tous les champs sont remplis
        existing_combinations.add(tuple(row))

# Saisie manuelle
nom = input("Entrez le nom : ").strip()
age = input("Entrez l'âge : ").strip()
ville = input("Entrez la ville : ").strip()

# Tenter de convertir l'âge en entier
try:
    age = int(age)
except ValueError:
    print("Âge invalide. Veuillez entrer un nombre.")
    exit()

# Créer une combinaison
new_data = (nom, age, ville)

# Vérifier si la combinaison existe déjà
if new_data in existing_combinations:
    print("Cette personne existe déjà dans le fichier. Aucune donnée ajoutée.")
else:
    # Ajouter les en-têtes si la feuille est vide
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
        headers = ['Nom', 'Âge', 'Ville']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

    # Ajouter la nouvelle ligne
    sheet.append(new_data)

    # Sauvegarder
    book.save(file_path)
    print("Données ajoutées avec succès.")
