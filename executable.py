# -*- coding: utf-8 -*-
"""
Created on Mon Apr 21 17:18:16 2025

@author: abouf
"""

import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os




# nom du fichier excel
fichier_excel = 'donneees.xlsx'

# creer le fichier s'il n'existe pas 
if not os.path_exists(fichier_excel):
    wb = openpyxl.Workbook()
    ws = wb.active()
    ws.append(["ID","Nom","Prenom","Email"])
    wb.save(fichier_excel)
    
# fonction pou charger les donnees dans le tableau

def charger_donnees():
    for item in tree.get_children():
        tree.delete(item)
    wb = openpyxl.load_workbook(fichier_excel)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)
    ws.close()
    
# fonction pour inserer les donnees
def ajouter():
    nom = entry_nom.get()
    prenom = entry_prenom.get()
    email = entry_email.get()

    if not nom or not prenom or not email:
        messagebox.showwarning("champs vides","veuillez remplir tous les champs")
        return
    wb.openpyxl.load_workbook(fichier_excel)
    ws.active
    next_id = ws.max_row
    ws.append([next_id, nom, prenom, email])
    wb.save(fichier_excel)
    
    charger_donnees()
    messagebox.showinfo("Succès","Enregistrement effectué avec succès")


# fonction pour remplir les champs depuis les treeVew

def selection(event):
    selected_item = tree.selection()
    if selected_item:
        values = tree.item(selected_item)
    ["Values"]
    
























# Interface graphique
root = tk.Tk()
root.title("CRUD avec tkinter")
root.geometry("700x500")

# Les champs de saisis
tk.Label(root, text="ID").grid(row=0, column=0, padx=5, pady=5)
entry_id = tk.Entry(root)
entry_id.grid(row=0, column=1, padx=5, pady=5)
entry_id.config(state='readonly')


tk.Label(root, text="Nom").grid(row=1, column=0, padx=5, pady=5)
entry_nom = tk.Entry(root)
entry_nom.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Prenom").grid(row=2, column=0, padx=5, pady=5)
entry_prenom = tk.Entry(root)
entry_prenom.grid(row=2, column=1, padx=5, pady=5)

tk.Label(root, text="Email").grid(row=3, column=0, padx=5, pady=5)
entry_email = tk.Entry(root)
entry_email.grid(row=3, column=1, padx=5, pady=5)

# Les bouttons 
tk.Button(root, text="Ajouter", command=ajouter).grid(row=4, column=0, pady=10)
tk.Button(root, text="Modifier", command=modifier).grid(row=4, column=1, pady=10)
tk.Button(root, text="Supprimer", command=supprimer).grid(row=4, column=2, pady=10)









# Le tableau pour la liste
tree = ttk.Treeview(root, columns=("ID","Nom","Prenom","Email"), show="headings")
tree.heading("ID", text="ID")
tree.heading("Nom", text="Nom")
tree.heading("Prenom", text="Prenom")
tree.heading("Email", text="Email")
tree.grid(row=5, column=0, columnspan=3, padx=10, pady=10)
tree.bind("TreevewSelect", selection)


root.mainloop()



























