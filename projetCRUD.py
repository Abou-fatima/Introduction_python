# -*- coding: utf-8 -*-
"""
Created on Mon May  5 14:18:25 2025

@author: LAMA MEDIA
"""

import tkinter as tk
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox


fenetre=tk.Tk()
fenetre.title("Enregistrement personnel")



fenetre.geometry("600x500+400+200")

libMatricule=tk.Label(fenetre,text="Matricule")
libMatricule.grid(row=0,column=0,padx=10,pady=10)
entryMatricule=tk.Entry(fenetre)
entryMatricule.grid(row=0,column=1,padx=10,pady=10)

libNom=tk.Label(fenetre,text="Nom")
libNom.grid(row=1,column=0,padx=10,pady=10)
entryNom=tk.Entry(fenetre)
entryNom.grid(row=1,column=1,padx=10,pady=10)

libAge=tk.Label(fenetre,text="Age")
libAge.grid(row=2,column=0,padx=10,pady=10)
entryAge=tk.Entry(fenetre)
entryAge.grid(row=2,column=1,padx=10,pady=10)

libVille=tk.Label(fenetre,text="Ville")
libVille.grid(row=3,column=0,padx=10,pady=10)
entryVille=tk.Entry(fenetre)
entryVille.grid(row=3,column=1,padx=10,pady=10)

entryRecherche=tk.Entry(fenetre)
entryRecherche.grid(row=4,column=6,padx=10,pady=10)


#Fonction d'insertion des données

def insertionDonnee():
    
    matricule=entryMatricule.get()
    nom=entryNom.get()
    age=entryAge.get()
    ville=entryVille.get()
    

    # Vérification des champs

    if not matricule or not nom or not age or not ville:
        messagebox.showerror("Erreur","Veuillez remplir tous les champs")
        return

    if not age.isdigit():
        messagebox.showerror("Veuillez donner un age valide")
        return

    data_to_add={
        'Matricule':[matricule],
        'Nom':[nom],
        'Age':[age],
        'Ville':[ville]
    }

    df_to_add=pd.DataFrame(data_to_add)
    file_path="C:/TPPYTHON/data.xlsx"

    try:
        book=load_workbook(file_path)
        print("Fichier chargé avec succès")

    except FileNotFoundError:
        print(f"Erreur: le fichier {file_path} n'existe pas")

    # Vérifier si la feuille existe, sinon créer une nouvelle feuille
    sheet_name = 'Feuille1'  # Remplacez par le nom de votre feuille
    
    # Si la feuille existe déjà, on la sélectionne
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
    else:
        # Si la feuille n'existe pas, on la crée
        sheet = book.create_sheet(sheet_name)
    
    # Ajouter les en-têtes si la feuille est vide
    if sheet.max_row == 1:
        for col_num, column_title in enumerate(df_to_add.columns, 1):
            sheet.cell(row=1, column=col_num, value=column_title)
    
    # Ajouter les données à la feuille existante après la dernière ligne
    for row in df_to_add.itertuples(index=False, name=None):
        sheet.append(row)
    
    # Sauvegarder le fichier Excel avec les nouvelles données
    book.save(file_path)
    messagebox.showinfo("Info","Ajout effectué avec succès")

    viderChamps()
    
    print(f"Les nouvelles données ont été ajoutées et le fichier a été sauvegardé dans {file_path}.")

#Fonction pour la recherche
def rechercher():
    rechercheval=entryRecherche.get().strip()

    try:
        data=pd.read_excel("C:/TPPYTHON/data.xlsx",sheet_name="Feuille1")
        print(data)
        dataRecherche=data[data['Matricule'].astype(str).str.strip() == rechercheval.strip()]
        if not dataRecherche.empty:
            matricule1=dataRecherche['Matricule'].iloc[0]
            nom1=dataRecherche['Nom'].iloc[0]
            age1=dataRecherche['Age'].iloc[0]
            ville1=dataRecherche['Ville'].iloc[0]

            entryMatricule.delete(0,tk.END)
            entryMatricule.insert(0,matricule1)

            entryNom.delete(0,tk.END)
            entryNom.insert(0,nom1)

            entryAge.delete(0,tk.END)
            entryAge.insert(0,age1)

            entryVille.delete(0,tk.END)
            entryVille.insert(0,ville1)
            
         #   messagebox.showinfo("Recherche","Trouvé")
            global ligne_trouve
            ligne_trouve=dataRecherche.iloc[0]  
        else:
            messagebox.showinfo("Recherche","Pas trouvé")
            viderChamps()
        print(data['Matricule'])

        print(dataRecherche)

    except:
        messagebox.showerror("Error","Aucune données saisie")

    
        
    

#Modification des données
def modifier_donnees():
    if not ligne_trouve.empty:
      
        nouveau_matricule=entryMatricule.get()
        nouveau_nom=entryNom.get()
        nouveau_age=entryAge.get()
        nouvelle_ville=entryVille.get()
        if not nouveau_matricule or not nouveau_nom or not nouveau_age or not nouvelle_ville:
            messagebox.showinfo("Attention","Tous les champs sont obligatoire")
            return 
        try:
            df=pd.read_excel("C:/TPPYTHON/data.xlsx",sheet_name="Feuille1")
       
            df.loc[df['Matricule']==ligne_trouve['Matricule'],'Matricule']=nouveau_matricule
            df.loc[df["Nom"]==ligne_trouve["Nom"],"Nom"]=nouveau_nom
            df.loc[df["Age"]==ligne_trouve["Age"],"Age"]=nouveau_age
            df.loc[df["Ville"]==ligne_trouve["Ville"],"Ville"]=nouvelle_ville
             
            df.to_excel("C:/TPPYTHON/data.xlsx",sheet_name="Feuille1",index=False)
            messagebox.showinfo("Success","Données modifiée avec succès")
            viderChamps()
        except Exception as e:
            messagebox.showerror("Erreur",f"Erreur lors de la modification: {str(e)}")
        else:
            messagebox.showarring("Attention","Aucun resultat trouvé à modifier")
    


def supprimer():
   if not ligne_trouve.empty:
       response=messagebox.askyesno("Suppression","Voulez-vous supprimer l'enregistrement ?")
       if response:
           try:
               df=pd.read_excel("C:/TPPYTHON/data.xlsx",sheet_name="Feuille1")
               df=df[df["Matricule"]!=ligne_trouve["Matricule"]]
               df.to_excel("C:/TPPYTHON/data.xlsx",sheet_name="Feuille1",index=False)
               messagebox.showinfo("Info","Suppression effectué avec succès")
               #viderChamps()
               entryRecherche.delete(0,tk.END)
           except Exception as e:
               messagebox.showwarning("Erreur",f"Erreur de suppression {e}")
   else:
       messagebox.showinfo("Suppression","Aucune donnéé à supprimer")
            
          

def viderChamps():
    entryAge.delete(0,tk.END)
    entryMatricule.delete(0,tk.END)
    entryNom.delete(0,tk.END)
    entryVille.delete(0,tk.END)





btnAjout=tk.Button(fenetre,text="Valider",command=insertionDonnee)
btnAjout.grid(row=4,column=0,padx=10,pady=10)

btnRecherche=tk.Button(fenetre,text="Rechercher",command=rechercher)
btnRecherche.grid(row=4,column=10,padx=10,pady=10)

btnModification=tk.Button(fenetre,text="Modifier",command=modifier_donnees)
btnModification.grid(row=4,column=1,padx=10,pady=10)

btnSuppression=tk.Button(fenetre,text="Supprimer",command=supprimer)
btnSuppression.grid(row=4,column=2,padx=10,pady=10)
fenetre.mainloop()
